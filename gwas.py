import pandas as pd
import openpyxl as xl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.styles.colors import RED
import numpy as np
from sklearn.decomposition import PCA
import statsmodels.api as sm

#read input file into openpyxl
input_file = input("Please input excel file to be analyzed: ")
input_wb = xl.load_workbook(input_file)
input_sheet = input_wb['Sheet1']

#collect all line ID's, including duplicates
id_list = []
for row in range(2, input_sheet.max_row + 1):
    line_id_cell = input_sheet.cell(row, 1)
    id_list.append(line_id_cell.value)

#read genotype matrix into pandas and remove lines not present in the input file
matrix_df = pd.read_excel('genotype_matrix.xlsx')
matrix_df = matrix_df[matrix_df.iloc[:, 0].isin(id_list)]
matrix_df = matrix_df.iloc[:, 1:]
matrix_df = matrix_df.reset_index(drop=True)

#remove any columns from genotype matrix that contain only 0's or only 1's
matrix_df = matrix_df.loc[:, (matrix_df != 0).any(axis=0)]
matrix_df = matrix_df.loc[:, (matrix_df != 1).any(axis=0)]

#read kinship matrix into pandas
kinship = pd.read_excel('kinship_matrix.xlsx')

#check which input lines have genotype data (indirectly removes duplicates)
matched_id_list = []
for id in kinship.columns:
    if id in id_list:
        matched_id_list.append(id)
#create a list of common ID's between genotype matrix and input excel file
id_list = matched_id_list

#create ID column of all 1135 accessions
id_column = kinship.iloc[:, 0]
#remove ID column from kinship matrix
kinship = kinship.iloc[:, 1:]

#determine first three principal components from kinship matrix
pca = PCA(n_components=3)
principalComponents = pca.fit_transform(kinship)
principalDf = pd.DataFrame(data = principalComponents, columns = ['PC1', 'PC2', 'PC3'])
#connect ID's with corresponding principal components
principalDf = pd.concat([id_column, principalDf], axis=1, sort=False)
#remove principal components for ID's not in input file
principalDf = principalDf[principalDf.iloc[:, 0].isin(id_list)]
#remove only ID column
principalDf = principalDf.iloc[:, 1:]
principalDf = principalDf.reset_index(drop=True)


#read the input excel file into pandas
input_df = pd.read_excel(input_file)
#average the phenotype data of duplicate lines
input_df = input_df.groupby(input_df.columns[0], as_index=False).mean()
#remove lines that do not have genotype data
input_df = input_df[input_df.iloc[:, 0].isin(id_list)]
#sort dataframe by ID number to allign with corresponding principal components
input_df = input_df.sort_values(by=input_df.columns[0])
input_df = input_df.reset_index(drop=True)
#create column of phenotype values
phenotype = input_df.iloc[:, 1]
#create a data frame with the input phenotype, PC1, PC2, and PC3 as covariates
X = pd.concat([phenotype, principalDf], axis=1, sort=False)
#add a constant to the dataframe
X = sm.add_constant(X)

#prep a Results sheet of input excel file for beta and p value outputs
#and delete the old Results sheet if one exists
try:
    input_wb.remove(input_wb['Results'])
except:
    pass
input_wb.create_sheet('Results')
results_sheet = input_wb['Results']
#add column names (B coef and p value)
b_title_cell = results_sheet.cell(1, 2)
b_title_cell.value = 'B coef'
p_title_cell = results_sheet.cell(1, 3)
p_title_cell.value = 'p value'
#add row names (gene ATG's)
counter = 2
for column_name in matrix_df.columns:
    gene_cell = results_sheet.cell(counter, 1)
    gene_cell.value = column_name
    counter += 1

#create a logistic regression model that is
#gene ~ constant + phenotype + PC1 + PC2 + PC3 for each of the (up to) 2088 genes
number_of_genes = len(matrix_df.columns)
for column in range(0, number_of_genes):
    y = matrix_df.iloc[:, column]
    test = sm.Logit(y, X)
    #use a try block to fit the model because it sometimes fails for low number of lines
    try:
        test = test.fit(disp=0)
    except:
        continue
    #collect B and p values of phenotype parameter in each gene model
    #and write to Results sheet of input excel file
    b = test.params[1]
    b_cell = results_sheet.cell(column + 2, 2)
    b_cell.value = b
    p = test.pvalues[1]
    p_cell = results_sheet.cell(column + 2, 3)
    p_cell.value = p
    #color p value cell based on level of significance:
    #no fill = not significant
    #orange = significant
    #red = highly significant
    if p < .001/number_of_genes:
        p_cell.fill = PatternFill(start_color=RED, end_color=RED, fill_type = 'solid')
    elif p < .05/number_of_genes:
        p_cell.fill = PatternFill(start_color='FFBB00', end_color='FFBB00', fill_type = 'solid')

#display number of unique lines with genotype data to user in terminal
print('Genotype data present for ' + str(len(X.index)) + ' of the input lines.')

#save the Results sheet to the input excel file
input_wb.save(input_file)
